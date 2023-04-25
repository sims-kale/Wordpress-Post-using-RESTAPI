import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
import os
import json
import jwt
import openpyxl
import json
from datetime import datetime, timedelta
import tempfile


def get_jwt_token(username, password):
    url = 'https://newbuildhomes.org/wp-json/jwt-auth/v1/token'
    params = {
        'username': username,
        'password': password
    }
    headers = {
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    response = requests.post(url=url, headers=headers, params=params)
    if response.ok:
        response_data = json.loads(response.text)
        return response_data['token']
    else:
        raise Exception(
            f"Failed to get JWT token. Status code: {response.status_code}. Response body: {response.text}")


def City(jwt_token, city_name):
    city_url = ' https://newbuildhomes.org/wp-json/wp/v2/property_city'

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    new_city = {
        'name': city_name
    }

    response = requests.post(url=city_url, headers=headers, json=new_city)
    city = json.loads(response.text)
    if response.ok:
        print("Property City added successfully")
        city_id = city['id']
        print(city_id)
        return city_id
    else:
        city_id = city['data']['term_id']
        print(city_id)
        return city_id


def Type(jwt_token, typestring):
    type_url = 'https://newbuildhomes.org/wp-json/wp/v2/property_type'
    print(type_url)
    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    new_type = {
        'name': typestring
    }
    response = requests.post(url=type_url, headers=headers, json=new_type)
    # print(response.text)
    type = json.loads(response.text)
    if response.ok:
        print("Property Type added successfully")
        type_id = type['id']
        print(type_id)
        return type_id
    else:
        type_id = type['data']['term_id']
        print(type_id)
        return type_id


def Media(jwt_token, imageurl):
    print(imageurl)
    media_url = "https://newbuildhomes.org/wp-json/wp/v2/media"

    raw = requests.get(imageurl).content

    file = tempfile.NamedTemporaryFile(delete=False, mode="wb", suffix=".jpg")
    filename = file.name
    print(filename)

    with file as img:
        img.write(raw)

    fileName = os.path.basename(filename)
    multipart_data = MultipartEncoder(
        fields={
            # a file upload field
            'file': (fileName, open(filename, 'rb'), 'image/jpeg'),
            # plain text fields
            'alt_text': 'alt test',
            'caption': 'caption test',
            'description': 'description test',
        }
    )

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': multipart_data.content_type,
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Content-Disposition': 'attachment; filename="1.1.jpeg"',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
    }

    response = requests.post(
        url=media_url, headers=headers, data=multipart_data)
    
    # print(response.text)

    media = json.loads(response.text)
    
    # print(media)
    
    if response.ok:
        print("Property images added successfully")
        media_id = media['id']
        # res = requests.post(
        #     'https://newbuildhomes.org/wp-json/wp/v2/properties/18225',
        #     headers=headers,
        #     json=[media_id]
        # )
        # print(res.text)
        # print(media_id)
        return (media_id)
    else:
        media_id = media['data']['term_id']
        print(media_id)
        return media_id


def create_post(jwt_token, post_url, title, city_id, type_id, media_id):

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
    }
    # for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, values_only=True):
    post_data = {
        'title': title,
        'fave_property_price': 12345,
        'status': 'publish',
        'property_city': city_id,
        'property_type': type_id,
        'featured_media': media_id

    }
    response = requests.post(url=post_url, headers=headers, json=post_data)
    print
    if response.ok:
        print('Post created successfully!')
    else:
        print(
            f'Error creating post. Status code: {response.status_code}. Response body: {response.text}')


def main():
    username = 'Muktesh'
    password = 'pNku V9CJ WvSQ 5jwZ Ip1S gPbV'
    jwt_token = get_jwt_token(username, password)
    post_url = 'https://newbuildhomes.org/wp-json/wp/v2/properties'

    wb = openpyxl.load_workbook('Property.xlsx')
    ws = wb['extraction results']

    for row in ws.iter_rows(min_row=9, max_row=15, min_col=1, values_only=True):
        print(row)
        title= row[1]
        city_name = row[2].split(',')[0]
        typestring = row[3]
        arr_images = row[6].replace("]", "").replace(
            "[", "").replace("'", "").split(",")
        print(arr_images)
        print('city_name', city_name)
        city = City(jwt_token, city_name)
        type = Type(jwt_token, typestring)
        for image in arr_images:
            media = Media(jwt_token, image)
        create_post(jwt_token, post_url, title, city, type, media)


if __name__ == '__main__':
    main()
