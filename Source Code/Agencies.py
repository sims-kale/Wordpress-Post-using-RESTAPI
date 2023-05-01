import tempfile
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
import os
import json
import jwt
import openpyxl
import json
import logging

logging.basicConfig(filename='Agency_Logs.log', level=logging.INFO)


def get_jwt_token(username, password):
    url = 'https://newbuildhomes.org/wp-json/jwt-auth/v1/token'
    params = {
        'username': username,
        'password': password
    }
    headers = {
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    response = requests.post(url=url, headers=headers, params=params)
    if response.ok:
        response_data = json.loads(response.text)
        return response_data['token']
    else:
        logging.exception(
            f"Failed to get JWT token. Status code: {response.status_code}. Response body: {response.text}")


def Media(jwt_token, image):
    if not image:
        logging.info("Image cell is empty, skipping...")
        return None
    print(image)

    image_path = os.path.join("D:/relu/Wordpress data/Agency_images", image)
    print(image_path)

    if not os.path.exists(image_path):
        logging.warning(f"Image file {image} does not exist, skipping...")
        return None

    media_url = "https://newbuildhomes.org/wp-json/wp/v2/media"

    with open(image_path, 'rb') as f:
        raw = f.read()
    file = tempfile.NamedTemporaryFile(delete=False, mode="wb", suffix=".png")
    filename = file.name

    with file as img:
        img.write(raw)

    multipart_data = MultipartEncoder(
        fields={
            # a file upload field

            'file': (image, open(filename, 'rb'), 'image/png'),
            # plain text fields
            'alt_text': 'alt test',
            'caption': 'caption test',
            'description': 'description test',
        }
    )

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': multipart_data.content_type,
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        # 'Content-Disposition': 'attachment; filename="1.1.jpeg"',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
    }

    response = requests.post(
        url=media_url, headers=headers, data=multipart_data)

    media = json.loads(response.text)

    # print(media)

    if response.ok:

        media_id = media['id']
        logging.info("Property images added successfully " + str(media_id))

        return (media_id)
    else:
        logging.warning(
            f"Failed to upload media '{image}' with status code {response.status_code}")


def create_post(jwt_token, post_url, title, media_id):
    print(title)
    if title is None:
        logging.info("Title is None, skipping...")
        return None
    logging.info("Agency Title: "+title)

    post_url = 'https://newbuildhomes.org/wp-json/wp/v2/agencies'

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
    }
    # for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, values_only=True):
    post_data = {
        'title': title,
        'featured_media': media_id,
        'status': 'publish'


    }
    response = requests.post(url=post_url, headers=headers, json=post_data)
    post = json.loads(response.text)

    if response.ok:
        post_id = post['id']
        logging.info("New post Created: " + str(post_id))
        # edit_post_url = "https://newbuildhomes.org/wp-admin/post.php?post=" +str(post_id)+"&action=edit"
        # logging.info("Post URL: " + edit_post_url)
        logging.info(
            "--------------------------------------------------------------------------------------------")
        return post_id

    else:
        logging.exception(
            f'Error creating post. Status code: {response.status_code}. Response body: {response.text}')


def main():
    username = 'Muktesh'
    password = 'pNku V9CJ WvSQ 5jwZ Ip1S gPbV'
    jwt_token = get_jwt_token(username, password)
    post_url = 'https://newbuildhomes.org/wp-json/wp/v2/properties'

    wb = openpyxl.load_workbook('Property.xlsx')
    ws = wb['Agencies']

    for row in ws.iter_rows(min_row=2, max_row=319, min_col=1, values_only=True):
        print(row)
        title = row[0]
        image = row[1]
        media = Media(jwt_token, image)
        create_post(jwt_token, post_url, title, media)


if __name__ == '__main__':
    main()
