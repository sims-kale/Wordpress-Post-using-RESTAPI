import re
import tempfile
import requests
from requests_toolbelt.multipart.encoder import MultipartEncoder
import os
import json
import jwt
import openpyxl
import json
import logging
from datetime import datetime, timedelta


logging.basicConfig(filename='Source Code/Posts.log', level=logging.INFO)


def get_jwt_token(username, password):
    url = 'https://newbuildhomes.org/wp-json/jwt-auth/v1/token'
    params = {
        'username': username,
        'password': password
    }
    headers = {
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    response = requests.post(url=url, headers=headers, params=params)
    if response.ok:
        response_data = json.loads(response.text)
        return response_data['token']
    else:
        logging.exception(
            f"Failed to get JWT token. Status code: {response.status_code}. Response body: {response.text}")


def City(jwt_token, city_name):
    if city_name is None:
        logging.info("City name is None, skipping...")
        return None
    logging.info("Property City: " + city_name)

    city_url = ' https://newbuildhomes.org/wp-json/wp/v2/property_city'

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    new_city = {
        'name': city_name
    }

    response = requests.post(url=city_url, headers=headers, json=new_city)
    city = json.loads(response.text)
    if response.ok:
        city_id = city['id']
        logging.info("Property City added successfully " + str(city_id))

        return city_id
    else:
        city_id = city['data']['term_id']
        logging.info("City already existed: " + str(city_id))
        return city_id
    
def Area(jwt_token, area_name):
    if area_name is None:
        logging.info("Area name is None, skipping...")
        return None
    logging.info("Property Area: " + area_name)

    area_url = ' https://newbuildhomes.org/wp-json/wp/v2/property_area'

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }
    new_area = {
        'name': area_name
    }

    response = requests.post(url=area_url, headers=headers, json=new_area)
    area = json.loads(response.text)
    if response.ok:
        area_id = area['id']
        logging.info("Property area added successfully " + str(area_id))

        return area_id
    else:
        area_id = area['data']['term_id']
        logging.info("area already existed: " + str(area_id))
        return area_id


def Type(jwt_token, typestring):
    
    if typestring is None:
        logging.info("Typestring is None, skipping...")
        return None
    
    logging.info("Property Type: " + typestring.strip())

    type_url = 'https://newbuildhomes.org/wp-json/wp/v2/property_type'

    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': '*/*',
        'User-Agent': 'Mozilla/5.0'
    }


    types = []
#1, 2 & 3 bedroom apartments and 3, 4 & 5 bedroom houses
    andsplit = typestring.split("and")
    # print(andsplit)
    for andstring in andsplit:
        parent = str(andstring).strip().split(" ")[-1]
        numbers = re.findall(r'\d+', andstring)
        new_type = {
        'name': str(parent)
        }
        # request to create parent
        response = requests.post(url=type_url, headers=headers, json=new_type)
        type = json.loads(response.text)
        # print(parent)
        if response.ok:
            type_id = type['id']
            logging.info("Property Type added successfully "+str(type_id))
            types.append(int(type_id))
        else:
            type_id = type['data']['term_id']
            logging.info("Property type already existed: " + str(type_id))
            types.append(int(type_id))         
        # request to create child
        for number in numbers:
            new_type = {
            'name': f'{number} bedrooms',
            'parent': type_id
            }
            response = requests.post(url=type_url, headers=headers, json=new_type)
            type = json.loads(response.text)
            if response.ok:
                child_type_id = type['id']
                logging.info("Property Type added successfully "+str(child_type_id))
                types.append(int(child_type_id))
            else:
                child_type_id = type['data']['term_id']
                logging.info("Property type already existed: " + str(child_type_id))
                types.append(int(child_type_id))
    
    # print(types)
    return types


def Media(jwt_token, rename):
    print(rename)

    if not rename:
        logging.info("Image URL is empty, skipping...")
        return None

    media_url = "https://newbuildhomes.org/wp-json/wp/v2/media"

    if os.path.isfile(rename):
        with open(rename, "rb") as f:
            raw = f.read()
            # file = tempfile.NamedTemporaryFile(delete=False, mode="wb", suffix=".jpeg")
            # filename = file.name
            # with file as img:
            #     img.write(raw)

        fileName = os.path.basename(rename)
        multipart_data = MultipartEncoder(
            fields={
                # a file upload field
                'file': (fileName, open(rename, 'rb'), 'image/jpeg'),
                # plain text fields
                'alt_text': 'alt test',
                'caption': 'caption test',
                'description': 'description test',
            }
    )

        headers = {
            'Authorization': f'Bearer {jwt_token}',
            'Content-Type': multipart_data.content_type,
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Content-Disposition': 'attachment; filename="1.1.jpeg"',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
        }

        response = requests.post(
            url=media_url, headers=headers, data=multipart_data)

        # print(response.text)

        media = json.loads(response.text)

        # print(media)

        if response.ok:

            media_id = media['id']
            logging.info("Property images added successfully "+ str(media_id))
            return (media_id)
        # else:
        #     media_id = media['data']['term_id']
        #     logging.info("Image already existed: "+media_id)
        #     return media_id


def create_post(jwt_token, post_url, title, city_id, area_id, type_id, media_id):
  
    headers = {
        'Authorization': f'Bearer {jwt_token}',
        'Content-Type': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/112.0'
    }
    post_data = {
        'title': title,
        'property_city': city_id,
        'property_area': area_id,
        'property_type': type_id,
        'featured_media': media_id,
        'status': 'publish',

    }
    response = requests.post(url=post_url, headers=headers, json=post_data)
    post = json.loads(response.text)

    if response.ok:
        post_id = post['id']
        logging.info("New post Created: " + str(post_id))
        update_url = "https://newbuildhomes.org/wp-admin/post.php?post=" +str(post_id)+"&action=edit"
        print(update_url)
        
        # ws.cell(row, column=9, value=update_url)
        # ws.rows += 1
        # wb.save('Property.xlsx')

        logging.info("Post URL: " + post_url)
        logging.info("-----------------------------------------------------------------------------------")
        return update_url

    else:
        logging.exception(
            f'Error creating post. Status code: {response.status_code}. Response body: {response.text}')


def main():
    username = 'Muktesh'
    password = 'pNku V9CJ WvSQ 5jwZ Ip1S gPbV'
    jwt_token = get_jwt_token(username, password)
    post_url = 'https://newbuildhomes.org/wp-json/wp/v2/properties'

    wb = openpyxl.load_workbook('Property.xlsx')
    ws = wb['extraction results']
    starting_row=2
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=1, values_only=True):

        title = row[1].split("in")[0]
        if title is None:
            logging.info("Title is None, skipping...")
            return None
        logging.info("Property Title: "+title)

        city_name = row[2].split(',')[0]
        area_name = row[2].split(',')[1]
        typestring = row[3]
        arr_images =  row[5].replace("]", "").replace("[", "").replace("'", "").split(",")[0]  #update Media
        new_name = arr_images.replace(" by", '').replace(" in", '')
        rename= new_name.replace(" ", "-")
        print(rename)

        city = City(jwt_token, city_name)
         
        area = Area(jwt_token, area_name)
        
        type = Type(jwt_token, typestring)

        media = Media(jwt_token, rename)

        update_url= create_post(jwt_token, post_url, title, city, area, type, media)
    
        
        ws.cell(row=starting_row, column=9, value=update_url)
        starting_row +=1
        

        wb.save('Property.xlsx')



if __name__ == '__main__':
    main()
